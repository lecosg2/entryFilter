import openpyxl

wb = openpyxl.load_workbook('Hillsdale-FR-24.xlsx')
ws = wb.active

startingLine = 11
endingLine = 18885
curLine = startingLine

wb2 = openpyxl.Workbook()
ws2 = wb2.active
refLine = 1


def isNewEntry(line):
    curLine = ws['A' + str(line)].value
    if not curLine:
        return 0

    count = 0

    for c in curLine:
        if c == '*':
            count += 1

    return count > 10


def findLastLine(fstLine):
    curLine = fstLine + 1

    while not isNewEntry(curLine):
        curLine += 1

    return curLine - 1


def keep(fstLine, lstLine):
    for line in range(fstLine, lstLine + 1):
        curLine = ws['A' + str(line)].value
        if not curLine:
            continue
        if curLine.find('Vac') != -1:
            return True
        if curLine.find('vac') != -1:
            return True

    return False


for line in range(1, 11):
        ws2['A' + str(refLine)] = ws['A' + str(line)].value
        refLine += 1  

while curLine < endingLine:
    fstLine = curLine
    lstLine = findLastLine(fstLine)
    if keep(fstLine, lstLine):
        print("copiando linhas " + str(fstLine) + " a " + str(lstLine))
        for line in range(fstLine, lstLine + 1):
            ws2['A' + str(refLine)] = ws['A' + str(line)].value
            refLine += 1
        print("linhas " + str(fstLine) + " a " + str(lstLine) + " copiadas")

    curLine = lstLine + 1

wb2.save('Hillsdale-FR-24-FILTERED.xlsx')
